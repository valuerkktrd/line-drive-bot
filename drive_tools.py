"""Google Drive tools สำหรับให้ Gemini เรียกใช้ (function calling)"""

import io
import json
import os

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

SCOPES = ["https://www.googleapis.com/auth/drive"]
ROOT_FOLDER_ID = os.environ["DRIVE_ROOT_FOLDER_ID"]
CLIENT_SECRET_FILE = os.environ.get("GOOGLE_CLIENT_SECRET_FILE", "client_secret.json")
TOKEN_FILE = os.environ.get("GOOGLE_TOKEN_FILE", "token.json")

# บนคลาวด์: ส่ง token ผ่าน env var GOOGLE_TOKEN_JSON แทนไฟล์ (ไฟล์ไม่ถูก commit)
if not os.path.exists(TOKEN_FILE) and os.environ.get("GOOGLE_TOKEN_JSON"):
    with open(TOKEN_FILE, "w", encoding="utf-8") as _f:
        _f.write(os.environ["GOOGLE_TOKEN_JSON"])

_service = None


def _creds():
    """OAuth ในนามบัญชีผู้ใช้ (service account อัปโหลดลง My Drive ไม่ได้แล้ว)"""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    elif not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
        creds = flow.run_local_server(port=0)
    with open(TOKEN_FILE, "w", encoding="utf-8") as f:
        f.write(creds.to_json())
    return creds


def drive():
    global _service
    if _service is None:
        _service = build("drive", "v3", credentials=_creds())
    return _service


FILE_FIELDS = "id, name, mimeType, modifiedTime, size, parents, webViewLink"


def _fmt(files: list[dict]) -> str:
    if not files:
        return "ไม่พบไฟล์"
    return json.dumps(files, ensure_ascii=False, indent=1)


def search_files(query: str) -> str:
    """ค้นหาไฟล์/โฟลเดอร์ใน Google Drive จากชื่อ (ค้นแบบ contains)

    Args:
        query: คำค้นหาในชื่อไฟล์ เช่น "ราคาประเมิน" หรือ "น้ำพอง"
    """
    q = query.replace("'", "\\'")
    res = drive().files().list(
        q=f"name contains '{q}' and trashed=false",
        fields=f"files({FILE_FIELDS})",
        pageSize=20,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    return _fmt(res.get("files", []))


def list_folder(folder_id: str = "") -> str:
    """แสดงรายการไฟล์ในโฟลเดอร์ (ถ้าไม่ระบุ folder_id จะใช้โฟลเดอร์หลักของบอท)

    Args:
        folder_id: ID ของโฟลเดอร์ที่ต้องการดู (เว้นว่าง = โฟลเดอร์หลัก)
    """
    fid = folder_id or ROOT_FOLDER_ID
    res = drive().files().list(
        q=f"'{fid}' in parents and trashed=false",
        fields=f"files({FILE_FIELDS})",
        pageSize=50,
        orderBy="folder,name",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    return _fmt(res.get("files", []))


def create_folder(name: str, parent_id: str = "") -> str:
    """สร้างโฟลเดอร์ใหม่

    Args:
        name: ชื่อโฟลเดอร์
        parent_id: ID โฟลเดอร์แม่ (เว้นว่าง = สร้างในโฟลเดอร์หลักของบอท)
    """
    meta = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id or ROOT_FOLDER_ID],
    }
    f = drive().files().create(
        body=meta, fields=FILE_FIELDS, supportsAllDrives=True
    ).execute()
    return _fmt([f])


def move_file(file_id: str, new_parent_id: str) -> str:
    """ย้ายไฟล์/โฟลเดอร์ไปยังโฟลเดอร์ปลายทาง

    Args:
        file_id: ID ของไฟล์ที่จะย้าย
        new_parent_id: ID ของโฟลเดอร์ปลายทาง
    """
    f = drive().files().get(
        fileId=file_id, fields="parents", supportsAllDrives=True
    ).execute()
    prev = ",".join(f.get("parents", []))
    f = drive().files().update(
        fileId=file_id,
        addParents=new_parent_id,
        removeParents=prev,
        fields=FILE_FIELDS,
        supportsAllDrives=True,
    ).execute()
    return _fmt([f])


def rename_file(file_id: str, new_name: str) -> str:
    """เปลี่ยนชื่อไฟล์/โฟลเดอร์

    Args:
        file_id: ID ของไฟล์
        new_name: ชื่อใหม่
    """
    f = drive().files().update(
        fileId=file_id, body={"name": new_name},
        fields=FILE_FIELDS, supportsAllDrives=True,
    ).execute()
    return _fmt([f])


def get_link(file_id: str) -> str:
    """ขอลิงก์เปิดดูไฟล์ (webViewLink)

    Args:
        file_id: ID ของไฟล์
    """
    f = drive().files().get(
        fileId=file_id, fields=FILE_FIELDS, supportsAllDrives=True
    ).execute()
    return _fmt([f])


MAX_READ_CHARS = 40000

# ---------- เครื่องมือวิเคราะห์ตาราง (ไฟล์ใหญ่หลักแสนแถว) ----------

_CACHE_DIR = os.environ.get("TABLE_CACHE_DIR", os.path.join(os.path.dirname(__file__), "_cache"))

TABLE_MIMES_XLSX = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
)


def _download_table(file_id: str):
    """ดาวน์โหลดไฟล์ตารางลง cache ครั้งเดียว คืน (path, ชนิด csv/xlsx, ชื่อไฟล์)"""
    meta = drive().files().get(
        fileId=file_id, fields="name, mimeType, md5Checksum, modifiedTime",
        supportsAllDrives=True,
    ).execute()
    mt, name = meta["mimeType"], meta["name"]
    ver = (meta.get("md5Checksum") or meta.get("modifiedTime", "")).replace(":", "")

    if mt == "application/vnd.google-apps.spreadsheet":
        kind = "csv"
    elif mt in TABLE_MIMES_XLSX:
        kind = "xlsx"
    elif mt.startswith("text/") or mt in ("application/csv",):
        kind = "csv"
    else:
        raise ValueError(f"ไฟล์ '{name}' ประเภท {mt} ไม่ใช่ตาราง (รองรับ Sheets/Excel/CSV)")

    os.makedirs(_CACHE_DIR, exist_ok=True)
    path = os.path.join(_CACHE_DIR, f"{file_id}_{ver}.{kind}")
    if not os.path.exists(path):
        if mt == "application/vnd.google-apps.spreadsheet":
            raw = drive().files().export(fileId=file_id, mimeType="text/csv").execute()
        else:
            raw = drive().files().get_media(fileId=file_id).execute()
        with open(path, "wb") as f:
            f.write(raw)
    return path, kind, name


def _load_df(file_id: str, sheet: str = "", columns: list | None = None):
    import pandas as pd

    path, kind, name = _download_table(file_id)
    if kind == "xlsx":
        df = pd.read_excel(path, sheet_name=(sheet or 0), engine="calamine",
                           usecols=columns or None)
    else:
        try:
            df = pd.read_csv(path, usecols=columns or None)
        except UnicodeDecodeError:
            df = pd.read_csv(path, usecols=columns or None, encoding="cp874")
    return df, name


def _apply_filter(df, filter_expr: str):
    if filter_expr:
        df = df.query(filter_expr, engine="python")
    return df


def file_stats(file_id: str, sheet: str = "") -> str:
    """ดูโครงสร้างไฟล์ตาราง: จำนวนแถว รายชื่อคอลัมน์ และตัวอย่าง 5 แถวแรก
    ใช้กับไฟล์ตารางขนาดใหญ่ก่อนเสมอ เพื่อรู้ชื่อคอลัมน์ก่อนจะ query/aggregate

    Args:
        file_id: ID ของไฟล์ (Sheets/Excel/CSV)
        sheet: ชื่อแผ่นงานใน Excel (เว้นว่าง = แผ่นแรก)
    """
    df, name = _load_df(file_id, sheet)
    cols = ", ".join(f"{c}({t})" for c, t in df.dtypes.astype(str).items())
    sample = df.head(5).to_string(index=False)
    return f"ไฟล์ '{name}': {len(df):,} แถว\nคอลัมน์: {cols}\nตัวอย่าง 5 แถวแรก:\n{sample}"


def query_file(file_id: str, filter_expr: str = "", columns: str = "",
               limit: int = 20, sheet: str = "") -> str:
    """ค้นหา/กรองแถวจากไฟล์ตารางขนาดใหญ่ คืนเฉพาะแถวที่ตรงเงื่อนไข (ไม่เกิน limit)

    Args:
        file_id: ID ของไฟล์ (Sheets/Excel/CSV)
        filter_expr: เงื่อนไขแบบ pandas query เช่น 'ราคา > 100000' หรือ
            'อำเภอ == "น้ำพอง" and ราคา >= 50000' หรือ 'ชื่อ.str.contains("แปลง")'
            (เว้นว่าง = เอาแถวแรกๆ)
        columns: รายชื่อคอลัมน์ที่ต้องการ คั่นด้วยจุลภาค (เว้นว่าง = ทุกคอลัมน์)
        limit: จำนวนแถวสูงสุดที่คืน (ค่าเริ่มต้น 20 สูงสุด 100)
        sheet: ชื่อแผ่นงานใน Excel (เว้นว่าง = แผ่นแรก)
    """
    df, name = _load_df(file_id, sheet)
    df = _apply_filter(df, filter_expr)
    total = len(df)
    if columns:
        want = [c.strip() for c in columns.split(",")]
        df = df[[c for c in want if c in df.columns]]
    out = df.head(max(1, min(int(limit), 100))).to_string(index=False)
    if len(out) > 8000:
        out = out[:8000] + "\n...(ตัดผลลัพธ์)"
    return f"พบ {total:,} แถวใน '{name}'\n{out}"


def aggregate_file(file_id: str, operation: str, column: str = "",
                   group_by: str = "", filter_expr: str = "", sheet: str = "") -> str:
    """คำนวณสรุปจากไฟล์ตารางขนาดใหญ่ (รวม เฉลี่ย นับ ต่ำสุด สูงสุด) แยกกลุ่มได้

    Args:
        file_id: ID ของไฟล์ (Sheets/Excel/CSV)
        operation: sum | mean | count | min | max | nunique
        column: คอลัมน์ตัวเลขที่จะคำนวณ (count/nunique เว้นว่างได้)
        group_by: คอลัมน์ที่ใช้แบ่งกลุ่ม เช่น 'อำเภอ' (เว้นว่าง = ทั้งไฟล์)
        filter_expr: กรองก่อนคำนวณ แบบ pandas query (เว้นว่าง = ทั้งหมด)
        sheet: ชื่อแผ่นงานใน Excel (เว้นว่าง = แผ่นแรก)
    """
    df, name = _load_df(file_id, sheet)
    df = _apply_filter(df, filter_expr)
    op = operation.strip().lower()
    if op not in ("sum", "mean", "count", "min", "max", "nunique"):
        return "operation ต้องเป็น: sum, mean, count, min, max, nunique"

    if group_by:
        g = df.groupby(group_by)
        series = g.size() if op == "count" and not column else getattr(g[column], op)()
        series = series.sort_values(ascending=False).head(100)
        body = series.to_string()
    elif op == "count" and not column:
        body = f"{len(df):,}"
    else:
        body = str(getattr(df[column], op)())

    cond = f" (เงื่อนไข: {filter_expr})" if filter_expr else ""
    return f"{op} ของ '{name}'{cond}:\n{body}"


def read_file(file_id: str) -> str:
    """อ่านเนื้อหาไฟล์เป็นข้อความ เพื่อนำไปสรุป ตอบคำถาม หรือคำนวณตัวเลข
    รองรับ: Google Sheets, Google Docs, Excel (.xlsx), CSV, TXT

    Args:
        file_id: ID ของไฟล์ที่จะอ่าน
    """
    meta = drive().files().get(
        fileId=file_id, fields="name, mimeType", supportsAllDrives=True
    ).execute()
    mt = meta["mimeType"]
    name = meta["name"]

    if mt == "application/vnd.google-apps.spreadsheet":
        raw = drive().files().export(fileId=file_id, mimeType="text/csv").execute()
        text = raw.decode("utf-8", errors="replace")
    elif mt == "application/vnd.google-apps.document":
        raw = drive().files().export(fileId=file_id, mimeType="text/plain").execute()
        text = raw.decode("utf-8", errors="replace")
    elif mt in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        from openpyxl import load_workbook

        raw = drive().files().get_media(fileId=file_id).execute()
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"### แผ่นงาน: {ws.title}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 500:
                    lines.append("...(ตัดที่ 500 แถว)")
                    break
                lines.append(",".join("" if c is None else str(c) for c in row))
        text = "\n".join(lines)
    elif mt.startswith("text/") or mt in ("application/json", "application/csv"):
        raw = drive().files().get_media(fileId=file_id).execute()
        text = raw.decode("utf-8", errors="replace")
    else:
        return f"ไฟล์ '{name}' เป็นประเภท {mt} — ยังอ่านเนื้อหาไม่ได้ (อ่านได้: Sheets, Docs, Excel, CSV, TXT)"

    if len(text) > MAX_READ_CHARS:
        text = text[:MAX_READ_CHARS] + f"\n...(ตัดเนื้อหา ไฟล์ยาวเกิน {MAX_READ_CHARS} ตัวอักษร)"
    return f"เนื้อหาไฟล์ '{name}':\n{text}"


def trash_file(file_id: str) -> str:
    """ย้ายไฟล์ไปถังขยะ (กู้คืนได้ใน 30 วัน) — เรียกใช้เฉพาะเมื่อผู้ใช้ยืนยันแล้วเท่านั้น

    Args:
        file_id: ID ของไฟล์ที่จะลบ
    """
    f = drive().files().update(
        fileId=file_id, body={"trashed": True},
        fields="id, name", supportsAllDrives=True,
    ).execute()
    return f"ย้ายไปถังขยะแล้ว: {f.get('name')} ({f.get('id')})"


ALL_TOOLS = [
    search_files,
    list_folder,
    create_folder,
    move_file,
    rename_file,
    get_link,
    read_file,
    file_stats,
    query_file,
    aggregate_file,
    trash_file,
]


def upload_bytes(data: bytes, filename: str, mime_type: str,
                 parent_id: str | None = None) -> dict:
    """อัปโหลดไฟล์ (ใช้จากฝั่ง webhook ตอนผู้ใช้ส่งไฟล์เข้ามาใน LINE)"""
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime_type, resumable=False)
    meta = {"name": filename, "parents": [parent_id or ROOT_FOLDER_ID]}
    return drive().files().create(
        body=meta, media_body=media, fields=FILE_FIELDS, supportsAllDrives=True
    ).execute()
